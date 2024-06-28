import openpyxl


def flatten_objects(obj, parent_key='', level=0, result=None, include_parent=False, include_containers=True):
    if result is None:
        result = []
    for k, v in obj.items():
        if include_parent:
            new_key = f'{parent_key}.{k}' if parent_key else k
        else:
            new_key = k
        if not isinstance(v, dict):
            result.append({'key': new_key, 'level': level, 'value': v})
        else:
            if (include_containers):
                result.append({'key': new_key, 'level': level, 'value': None})
            flatten_objects(v, new_key, level + 1, result,
                            include_parent, include_containers)
    return result


class ExcelManager():
    def __init__(self, filename):
        self.filename = filename
        self.wb = openpyxl.Workbook()
        self.last_row = 1
        self.save()

    def save(self):
        self.wb.save(self.filename)

    def add_row(self, row):
        if (row['N'] == 0):
            self.create_header(row)
        ws = self.wb.active
        cols = flatten_objects(row, include_containers=False)
        col_index = 0
        for col in cols:
            col_index += 1
            if col['key'] == 'clients':
                for client in col['value']:
                    name = 'Cli ' + str(client.get_id())
                    if client.node:  # If the client is being attended
                        name += ' - ' + str(client.node)
                    ws.cell(row=self.last_row, column=col_index, value=name)
                    col_index += 1
                    ws.cell(row=self.last_row, column=col_index,
                            value=client.status + ' - ' + client.service)
                    col_index += 1
                    ws.cell(row=self.last_row, column=col_index,
                            value=client.arrival_time)
                    col_index += 1
                continue
            ws.cell(row=self.last_row, column=col_index, value=col['value'])
        self.last_row += 1
        self.save()

    def create_header(self, starting_row):
        ws = self.wb.active
        cols = flatten_objects(starting_row)
        last_row_i = 1
        col_index = 0
        for col in cols:
            col_index += 1
            row_index = col['level'] + 1
            if last_row_i >= self.last_row:
                self.last_row += 1
            if row_index > last_row_i:
                col_index -= 1
            ws.cell(row=row_index, column=col_index, value=col['key'])
            last_row_i = row_index
