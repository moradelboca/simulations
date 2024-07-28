import openpyxl

EXCEL_PATH = './TP5/simulation.xlsx'


class SimulationPrinter():
    def __init__(self):
        self.eh = excelHandler
        self.add_rk_header()
        self.lock_rk = False

    def add_rk_row(self, values):
        if (not self.lock_rk):
            self.eh.add_rk_row(values)

    def add_rk_header(self):
        self.eh.add_rk_row(['x', 'y', 'k1', 'k2', 'k3', 'k4'])

    def get_values(self, row, result=None):
        if result is None:
            result = []
        for key, value in row.items():
            if isinstance(value, dict):
                self.get_values(value, result)
            elif isinstance(value, list) and key == 'clients':
                for cli in value:
                    result.append(cli.__str__())
            else:
                result.append(value)
        return result

    def get_header(self, row, result=None, level=0):
        # This just works, please dont ask how
        if result is None:
            result = [[]]
        first = False
        if len(result[level]) > 0 and result[level][-1] == '':
            first = True
        for key, value in row.items():
            for i in range(len(result)):
                if i == level:
                    if first:
                        result[i].pop()
                    result[i].append(key)
                else:
                    if not first:
                        result[i].append('')
            first = False
            if isinstance(value, dict):
                if level >= len(result) - 1:
                    result.append(['' for i in range(len(result[0]))])
                self.get_header(value, result, level + 1)
        return result

    def add_row(self, row):
        values = self.get_values(row)
        self.eh.add_sim_row(values)

    def add_header(self, row):
        header_rows = self.get_header(row)
        for header_row in header_rows:
            self.eh.add_sim_row(header_row)


class ExcelHandler():
    def __init__(self, filename):
        self.filename = filename
        self.wb = openpyxl.Workbook()
        # Remove default sheet
        self.wb.remove(self.wb.active)
        self.wb.create_sheet(title='Simulation')
        self.wb.create_sheet(title='Runge Kutta')
        self.ws = self.wb['Simulation']

    def add_sim_row(self, row):
        if self.ws.title != 'Simulation':
            self.ws = self.wb['Simulation']
        self.ws.append(row)
        self.save()

    def add_rk_row(self, values):
        if self.ws.title != 'Runge Kutta':
            self.ws = self.wb['Runge Kutta']
        self.ws.append(values)
        self.save()

    def save(self):
        self.wb.save(self.filename)


excelHandler = ExcelHandler(EXCEL_PATH)
simulationPrinter = SimulationPrinter()
